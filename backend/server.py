from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Form
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import logging
import bcrypt
import jwt
import secrets
from pydantic import BaseModel, Field, ConfigDict
from typing import Optional, List
from datetime import datetime, timezone, timedelta
import re
import math

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
JWT_ALGORITHM = "HS256"

def get_jwt_secret():
    return os.environ["JWT_SECRET"]

# Password utils
def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode("utf-8"), salt).decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

# Token utils
def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id, "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=60),
        "type": "access"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

# Auth dependency
async def get_current_user(request: Request) -> dict:
    auth_header = request.headers.get("Authorization", "")
    token = None
    if auth_header.startswith("Bearer "):
        token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Non authentifié")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Type de token invalide")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Utilisateur non trouvé")
        if user.get("active") is False:
            raise HTTPException(status_code=403, detail="Compte désactivé. Contactez l'administrateur.")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expiré")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalide")

# App setup
app = FastAPI(title="R2A Industrie - Gestion de Stock")
api_router = APIRouter(prefix="/api")

# Models
class LoginRequest(BaseModel):
    email: str
    password: str

class RegisterRequest(BaseModel):
    email: str
    password: str
    name: str
    role: Optional[str] = "employee"

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    created_at: Optional[str] = None

class ProductCreate(BaseModel):
    reference: str
    name: str
    category: str
    quantity: int = 0
    stock_minimum: int = 5
    purchase_price: float = 0.0
    sale_price: float = 0.0
    supplier: str = ""
    location: str = ""
    brand: str = ""
    description: str = ""
    state: str = "neuf"  # neuf / occasion / obsolete
    status: str = "en_stock"  # en_stock / rupture / sur_commande

class ProductUpdate(BaseModel):
    reference: Optional[str] = None
    name: Optional[str] = None
    category: Optional[str] = None
    quantity: Optional[int] = None
    stock_minimum: Optional[int] = None
    purchase_price: Optional[float] = None
    sale_price: Optional[float] = None
    supplier: Optional[str] = None
    location: Optional[str] = None
    brand: Optional[str] = None
    description: Optional[str] = None
    state: Optional[str] = None
    status: Optional[str] = None

# --- Accounting Models ---
class RevenueCreate(BaseModel):
    description: str
    amount: float
    category: str = "ventes"
    client_name: Optional[str] = ""
    invoice_ref: Optional[str] = ""
    payment_method: str = "virement"
    date: Optional[str] = None

class RevenueUpdate(BaseModel):
    description: Optional[str] = None
    amount: Optional[float] = None
    category: Optional[str] = None
    client_name: Optional[str] = None
    invoice_ref: Optional[str] = None
    payment_method: Optional[str] = None
    date: Optional[str] = None

class ExpenseCreate(BaseModel):
    description: str
    amount: float
    category: str = "achats_stock"
    supplier_name: Optional[str] = ""
    payment_method: str = "virement"
    date: Optional[str] = None

class ExpenseUpdate(BaseModel):
    description: Optional[str] = None
    amount: Optional[float] = None
    category: Optional[str] = None
    supplier_name: Optional[str] = None
    payment_method: Optional[str] = None
    date: Optional[str] = None

class InvoiceCreate(BaseModel):
    client_name: str
    client_email: Optional[str] = ""
    client_address: Optional[str] = ""
    items: list = []
    discount: float = 0
    notes: Optional[str] = ""
    due_date: Optional[str] = None

class InvoiceUpdate(BaseModel):
    client_name: Optional[str] = None
    client_email: Optional[str] = None
    client_address: Optional[str] = None
    items: Optional[list] = None
    discount: Optional[float] = None
    notes: Optional[str] = None
    due_date: Optional[str] = None
    status: Optional[str] = None

# --- Client Models ---
class ClientCreate(BaseModel):
    name: str
    phone: str = ""
    email: str = ""
    address: str = ""

class ClientUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None

# --- Sale Models ---
class SaleItemModel(BaseModel):
    product_id: str
    reference: str
    name: str
    quantity: int
    unit_price: float

class SaleCreate(BaseModel):
    client_id: Optional[str] = None
    client_name: str
    items: List[SaleItemModel]
    discount: float = 0
    notes: str = ""

# --- User Management Models ---
class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str = "employee"
    permissions: List[str] = []

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    active: Optional[bool] = None
    permissions: Optional[List[str]] = None
    password: Optional[str] = None

# ============ AUTH ROUTES ============

@api_router.post("/auth/login")
async def login(req: LoginRequest, request: Request):
    email = req.email.strip().lower()
    ip = request.client.host if request.client else "unknown"
    identifier = f"{ip}:{email}"
    attempt = await db.login_attempts.find_one({"identifier": identifier}, {"_id": 0})
    if attempt and attempt.get("count", 0) >= 5:
        lock_until = attempt.get("locked_until")
        if lock_until and datetime.now(timezone.utc) < lock_until:
            raise HTTPException(status_code=429, detail="Trop de tentatives. Réessayez dans 15 minutes.")
        else:
            await db.login_attempts.delete_one({"identifier": identifier})

    user = await db.users.find_one({"email": email})
    if not user or not verify_password(req.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$inc": {"count": 1}, "$set": {"locked_until": datetime.now(timezone.utc) + timedelta(minutes=15)}},
            upsert=True
        )
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")

    await db.login_attempts.delete_many({"identifier": identifier})
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    await log_activity(user_id, "login", "Connexion réussie")

    return {
        "id": user_id, "email": user["email"],
        "name": user.get("name", ""), "role": user.get("role", "employee"),
        "access_token": access_token, "refresh_token": refresh_token
    }

@api_router.post("/auth/register")
async def register(req: RegisterRequest):
    email = req.email.strip().lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Cet email est déjà utilisé")

    user_doc = {
        "email": email,
        "password_hash": hash_password(req.password),
        "name": req.name.strip(),
        "role": req.role if req.role in ["admin", "employee"] else "employee",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)

    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)

    return {
        "id": user_id, "email": email, "name": req.name.strip(), "role": user_doc["role"],
        "access_token": access_token, "refresh_token": refresh_token
    }

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return {
        "id": user["_id"], "email": user["email"],
        "name": user.get("name", ""), "role": user.get("role", "employee")
    }

@api_router.post("/auth/logout")
async def logout():
    return {"message": "Déconnecté"}

class RefreshRequest(BaseModel):
    refresh_token: str

@api_router.post("/auth/refresh")
async def refresh_token_endpoint(req: RefreshRequest):
    try:
        payload = jwt.decode(req.refresh_token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Type de token invalide")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Utilisateur non trouvé")
        user_id = str(user["_id"])
        new_access = create_access_token(user_id, user["email"])
        return {"access_token": new_access}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expiré")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalide")

# ============ DASHBOARD ROUTES ============

PRODUCT_CATEGORIES = [
    {"id": "automate", "name": "Automate", "icon": "Cpu"},
    {"id": "variateur", "name": "Variateur", "icon": "SlidersHorizontal"},
    {"id": "verin_pneumatique", "name": "Vérin Pneumatique", "icon": "Wind"},
    {"id": "vapeur", "name": "Vapeur", "icon": "Flame"},
    {"id": "relais_securite", "name": "Relais de sécurité + Capteur", "icon": "ShieldCheck"},
    {"id": "ecran_siemens", "name": "Ecran SIEMENS", "icon": "Monitor"},
    {"id": "hydrolique", "name": "Hydrolique", "icon": "Droplets"},
    {"id": "pneumatique", "name": "Pneumatique", "icon": "Gauge"},
    {"id": "encodeur_occasion", "name": "Encodeur occasion", "icon": "RotateCw"},
    {"id": "instrument", "name": "Instrument", "icon": "Radio"},
    {"id": "compteur", "name": "Compteur", "icon": "Hash"},
    {"id": "capteur", "name": "Capteur", "icon": "Radar"},
]

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(request: Request):
    await get_current_user(request)

    total_products = await db.products.count_documents({"archived": {"$ne": True}})
    low_stock = await db.products.count_documents({
        "archived": {"$ne": True},
        "$expr": {"$lte": ["$quantity", "$stock_minimum"]}
    })
    out_of_stock = await db.products.count_documents({
        "archived": {"$ne": True},
        "quantity": 0
    })

    # Monthly sales
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    monthly_sales_cursor = db.sales.aggregate([
        {"$match": {"created_at": {"$gte": month_start.isoformat()}, "archived": {"$ne": True}}},
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}, "count": {"$sum": 1}}}
    ])
    monthly_sales_result = await monthly_sales_cursor.to_list(1)
    monthly_sales = monthly_sales_result[0] if monthly_sales_result else {"total": 0, "count": 0}

    total_clients = await db.clients.count_documents({"archived": {"$ne": True}})

    # Category counts
    category_counts = {}
    for cat in PRODUCT_CATEGORIES:
        count = await db.products.count_documents({
            "category": cat["id"],
            "archived": {"$ne": True}
        })
        category_counts[cat["id"]] = count

    return {
        "total_products": total_products,
        "low_stock": low_stock,
        "out_of_stock": out_of_stock,
        "monthly_sales_amount": monthly_sales.get("total", 0),
        "monthly_sales_count": monthly_sales.get("count", 0),
        "total_clients": total_clients,
        "categories": PRODUCT_CATEGORIES,
        "category_counts": category_counts
    }

# ============ PRODUCTS ROUTES ============

def serialize_product(doc):
    """Convert MongoDB product document to JSON-safe dict."""
    return {
        "id": str(doc["_id"]),
        "reference": doc.get("reference", ""),
        "name": doc.get("name", ""),
        "category": doc.get("category", ""),
        "quantity": doc.get("quantity", 0),
        "stock_minimum": doc.get("stock_minimum", 5),
        "purchase_price": doc.get("purchase_price", 0),
        "sale_price": doc.get("sale_price", 0),
        "supplier": doc.get("supplier", ""),
        "location": doc.get("location", ""),
        "brand": doc.get("brand", ""),
        "description": doc.get("description", ""),
        "state": doc.get("state", "neuf"),
        "status": doc.get("status", "en_stock"),
        "archived": doc.get("archived", False),
        "archived_at": doc.get("archived_at"),
        "created_at": doc.get("created_at", ""),
        "updated_at": doc.get("updated_at", ""),
    }

@api_router.get("/products")
async def list_products(
    request: Request,
    category: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    archived: bool = False,
    sort_by: str = "created_at",
    sort_order: str = "desc",
):
    await get_current_user(request)

    query = {"archived": True} if archived else {"archived": {"$ne": True}}
    if category:
        query["category"] = category
    if search:
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query["$or"] = [
            {"reference": pattern},
            {"name": pattern},
            {"brand": pattern},
            {"supplier": pattern},
        ]

    sort_dir = -1 if sort_order == "desc" else 1
    skip = (page - 1) * limit

    total = await db.products.count_documents(query)
    cursor = db.products.find(query).sort(sort_by, sort_dir).skip(skip).limit(limit)
    products = [serialize_product(doc) async for doc in cursor]

    return {
        "products": products,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": math.ceil(total / limit) if limit > 0 else 0,
    }

@api_router.get("/products/{product_id}")
async def get_product(product_id: str, request: Request):
    await get_current_user(request)
    doc = await db.products.find_one({"_id": ObjectId(product_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Produit non trouvé")
    return serialize_product(doc)

@api_router.post("/products")
async def create_product(product: ProductCreate, request: Request):
    user = await get_current_user(request)
    ref = product.reference.strip()
    if not ref:
        raise HTTPException(status_code=400, detail="La référence est obligatoire")

    existing = await db.products.find_one({"reference": ref})
    if existing:
        raise HTTPException(status_code=400, detail=f"La référence '{ref}' existe déjà")

    now = datetime.now(timezone.utc).isoformat()
    doc = {
        **product.model_dump(),
        "reference": ref,
        "archived": False,
        "created_at": now,
        "updated_at": now,
    }
    result = await db.products.insert_one(doc)
    await log_activity(user["_id"], "product_create", f"Produit créé: {ref}")

    created = await db.products.find_one({"_id": result.inserted_id})
    return serialize_product(created)

@api_router.put("/products/{product_id}")
async def update_product(product_id: str, product: ProductUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.products.find_one({"_id": ObjectId(product_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Produit non trouvé")

    updates = {k: v for k, v in product.model_dump().items() if v is not None}
    if "reference" in updates:
        ref = updates["reference"].strip()
        dup = await db.products.find_one({"reference": ref, "_id": {"$ne": ObjectId(product_id)}})
        if dup:
            raise HTTPException(status_code=400, detail=f"La référence '{ref}' existe déjà")
        updates["reference"] = ref

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.products.update_one({"_id": ObjectId(product_id)}, {"$set": updates})
    await log_activity(user["_id"], "product_update", f"Produit modifié: {existing.get('reference')}")

    updated = await db.products.find_one({"_id": ObjectId(product_id)})
    return serialize_product(updated)

@api_router.delete("/products/{product_id}")
async def archive_product(product_id: str, request: Request):
    user = await get_current_user(request)
    existing = await db.products.find_one({"_id": ObjectId(product_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Produit non trouvé")

    await db.products.update_one(
        {"_id": ObjectId(product_id)},
        {"$set": {"archived": True, "archived_at": datetime.now(timezone.utc).isoformat()}}
    )
    await log_activity(user["_id"], "product_archive", f"Produit archivé: {existing.get('reference')}")
    return {"message": "Produit déplacé dans la corbeille"}

@api_router.post("/products/{product_id}/restore")
async def restore_product(product_id: str, request: Request):
    user = await get_current_user(request)
    existing = await db.products.find_one({"_id": ObjectId(product_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Produit non trouvé")

    await db.products.update_one(
        {"_id": ObjectId(product_id)},
        {"$set": {"archived": False}, "$unset": {"archived_at": ""}}
    )
    await log_activity(user["_id"], "product_restore", f"Produit restauré: {existing.get('reference')}")
    return {"message": "Produit restauré"}

@api_router.delete("/products/{product_id}/permanent")
async def delete_product_permanently(product_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Seul un administrateur peut supprimer définitivement")
    existing = await db.products.find_one({"_id": ObjectId(product_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Produit non trouvé")
    await db.products.delete_one({"_id": ObjectId(product_id)})
    await log_activity(user["_id"], "product_delete", f"Produit supprimé: {existing.get('reference')}")
    return {"message": "Produit supprimé définitivement"}

# ============ CLIENTS ROUTES ============

def serialize_client(doc):
    return {
        "id": str(doc["_id"]),
        "name": doc.get("name", ""),
        "phone": doc.get("phone", ""),
        "email": doc.get("email", ""),
        "address": doc.get("address", ""),
        "archived": doc.get("archived", False),
        "total_purchases": doc.get("total_purchases", 0),
        "created_at": doc.get("created_at", ""),
        "updated_at": doc.get("updated_at", ""),
    }

@api_router.get("/clients")
async def list_clients(request: Request, search: Optional[str] = None, page: int = 1, limit: int = 50, archived: bool = False):
    await get_current_user(request)
    query = {"archived": True} if archived else {"archived": {"$ne": True}}
    if search:
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query["$or"] = [{"name": pattern}, {"phone": pattern}, {"email": pattern}]
    total = await db.clients.count_documents(query)
    cursor = db.clients.find(query).sort("name", 1).skip((page - 1) * limit).limit(limit)
    items = [serialize_client(doc) async for doc in cursor]
    return {"items": items, "total": total, "page": page, "pages": math.ceil(total / limit) if limit > 0 else 0}

@api_router.post("/clients")
async def create_client(data: ClientCreate, request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    doc = {**data.model_dump(), "archived": False, "total_purchases": 0, "created_at": now, "updated_at": now}
    result = await db.clients.insert_one(doc)
    await log_activity(user["_id"], "client_create", f"Client créé: {data.name}")
    created = await db.clients.find_one({"_id": result.inserted_id})
    return serialize_client(created)

@api_router.put("/clients/{client_id}")
async def update_client(client_id: str, data: ClientUpdate, request: Request):
    user = await get_current_user(request)
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.clients.update_one({"_id": ObjectId(client_id)}, {"$set": updates})
    await log_activity(user["_id"], "client_update", f"Client modifié: {client_id}")
    updated = await db.clients.find_one({"_id": ObjectId(client_id)})
    if not updated:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    return serialize_client(updated)

@api_router.delete("/clients/{client_id}")
async def archive_client(client_id: str, request: Request):
    user = await get_current_user(request)
    await db.clients.update_one({"_id": ObjectId(client_id)}, {"$set": {"archived": True, "archived_at": datetime.now(timezone.utc).isoformat()}})
    await log_activity(user["_id"], "client_archive", f"Client archivé: {client_id}")
    return {"message": "Client déplacé dans la corbeille"}

@api_router.get("/clients/{client_id}/purchases")
async def client_purchases(client_id: str, request: Request):
    await get_current_user(request)
    cursor = db.sales.find({"client_id": client_id}).sort("created_at", -1).limit(100)
    sales = []
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        sales.append(doc)
    return {"items": sales}

# ============ SALES ROUTES ============

@api_router.get("/sales")
async def list_sales(request: Request, page: int = 1, limit: int = 50, month: Optional[str] = None, search: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if month:
        query["created_at"] = {"$regex": f"^{month}"}
    if search:
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query["$or"] = [{"client_name": pattern}, {"sale_number": pattern}]
    total = await db.sales.count_documents(query)
    cursor = db.sales.find(query).sort("created_at", -1).skip((page - 1) * limit).limit(limit)
    items = []
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        items.append(doc)
    return {"items": items, "total": total, "page": page, "pages": math.ceil(total / limit) if limit > 0 else 0}

@api_router.get("/sales/summary")
async def sales_summary(request: Request):
    await get_current_user(request)
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()

    # Monthly stats
    monthly = await db.sales.aggregate([
        {"$match": {"created_at": {"$gte": month_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}, "count": {"$sum": 1}}}
    ]).to_list(1)

    # Top products this month
    top_products = await db.sales.aggregate([
        {"$match": {"created_at": {"$gte": month_start}}},
        {"$unwind": "$items"},
        {"$group": {"_id": "$items.reference", "name": {"$first": "$items.name"}, "total_qty": {"$sum": "$items.quantity"}, "total_amount": {"$sum": {"$multiply": ["$items.quantity", "$items.unit_price"]}}}},
        {"$sort": {"total_qty": -1}},
        {"$limit": 10}
    ]).to_list(10)

    # Last 6 months
    months_data = []
    for i in range(6):
        m = now.month - i
        y = now.year
        if m <= 0:
            m += 12
            y -= 1
        ms = datetime(y, m, 1, tzinfo=timezone.utc).isoformat()
        me = datetime(y + 1, 1, 1, tzinfo=timezone.utc).isoformat() if m == 12 else datetime(y, m + 1, 1, tzinfo=timezone.utc).isoformat()
        r = await db.sales.aggregate([{"$match": {"created_at": {"$gte": ms, "$lt": me}}}, {"$group": {"_id": None, "total": {"$sum": "$total_amount"}, "count": {"$sum": 1}}}]).to_list(1)
        months_data.append({"month": f"{y}-{str(m).zfill(2)}", "total": r[0]["total"] if r else 0, "count": r[0]["count"] if r else 0})

    return {
        "month_total": monthly[0]["total"] if monthly else 0,
        "month_count": monthly[0]["count"] if monthly else 0,
        "top_products": top_products,
        "months": list(reversed(months_data)),
    }

@api_router.post("/sales")
async def create_sale(data: SaleCreate, request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    count = await db.sales.count_documents({})
    sale_number = f"VTE-{now.strftime('%Y%m')}-{str(count + 1).zfill(4)}"

    subtotal = sum(item.quantity * item.unit_price for item in data.items)
    total = subtotal - data.discount

    sale_doc = {
        "sale_number": sale_number,
        "client_id": data.client_id or "",
        "client_name": data.client_name,
        "items": [item.model_dump() for item in data.items],
        "subtotal": subtotal,
        "discount": data.discount,
        "total_amount": total,
        "notes": data.notes,
        "sold_by": user["_id"],
        "sold_by_name": user.get("name", ""),
        "created_at": now.isoformat(),
    }
    await db.sales.insert_one(sale_doc)

    # Update stock quantities and status
    for item in data.items:
        product = await db.products.find_one({"_id": ObjectId(item.product_id)})
        if product:
            new_qty = max(0, product.get("quantity", 0) - item.quantity)
            new_status = "rupture" if new_qty == 0 else "en_stock"
            await db.products.update_one(
                {"_id": ObjectId(item.product_id)},
                {"$set": {"quantity": new_qty, "status": new_status, "updated_at": now.isoformat()}}
            )

    # Update client total purchases
    if data.client_id:
        await db.clients.update_one({"_id": ObjectId(data.client_id)}, {"$inc": {"total_purchases": 1}})

    await log_activity(user["_id"], "sale_create", f"Vente {sale_number}: {data.client_name} - {total}€")
    return {**sale_doc, "_id": str(sale_doc.get("_id", ""))}

# ============ USER MANAGEMENT ROUTES ============

@api_router.get("/users")
async def list_users(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    cursor = db.users.find({}, {"password_hash": 0})
    items = []
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        items.append(doc)
    return {"items": items}

@api_router.post("/users")
async def create_user(data: UserCreate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    email = data.email.strip().lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Cet email existe déjà")
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "email": email, "password_hash": hash_password(data.password),
        "name": data.name, "role": data.role, "active": True,
        "permissions": data.permissions, "created_at": now,
    }
    result = await db.users.insert_one(doc)
    await log_activity(user["_id"], "user_create", f"Utilisateur créé: {email}")
    return {"id": str(result.inserted_id), "email": email, "name": data.name, "role": data.role}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, data: UserUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if "password" in updates:
        updates["password_hash"] = hash_password(updates.pop("password"))
    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": updates})
    await log_activity(user["_id"], "user_update", f"Utilisateur modifié: {user_id}")
    return {"message": "Utilisateur mis à jour"}

@api_router.delete("/users/{user_id}")
async def deactivate_user(user_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"active": False}})
    await log_activity(user["_id"], "user_deactivate", f"Utilisateur désactivé: {user_id}")
    return {"message": "Utilisateur désactivé"}

# ============ COMPANY SETTINGS ============

DEFAULT_COMPANY = {
    "name": "R2A Industrie",
    "address": "",
    "phone": "",
    "email": "",
    "rc": "",
    "nif": "",
    "nis": "",
    "ai": "",
}

@api_router.get("/settings/company")
async def get_company_settings(request: Request):
    await get_current_user(request)
    settings = await db.settings.find_one({"type": "company"}, {"_id": 0})
    return settings or DEFAULT_COMPANY

@api_router.put("/settings/company")
async def update_company_settings(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    body = await request.json()
    body["type"] = "company"
    await db.settings.update_one({"type": "company"}, {"$set": body}, upsert=True)
    await log_activity(user["_id"], "settings_update", "Paramètres entreprise modifiés")
    return {"message": "Paramètres enregistrés"}

# ============ DOCUMENTS (Facture + BL) ============

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib import colors
from starlette.responses import StreamingResponse

@api_router.get("/documents")
async def list_documents(request: Request, doc_type: Optional[str] = None, search: Optional[str] = None, page: int = 1, limit: int = 50):
    await get_current_user(request)
    query = {}
    if doc_type:
        query["doc_type"] = doc_type
    if search:
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query["$or"] = [{"doc_number": pattern}, {"client_name": pattern}]
    total = await db.documents.count_documents(query)
    cursor = db.documents.find(query).sort("created_at", -1).skip((page - 1) * limit).limit(limit)
    items = []
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        items.append(doc)
    return {"items": items, "total": total, "page": page, "pages": math.ceil(total / limit) if limit > 0 else 0}

@api_router.post("/documents")
async def create_document(request: Request):
    user = await get_current_user(request)
    body = await request.json()
    doc_type = body.get("doc_type", "facture")  # facture or bl
    now = datetime.now(timezone.utc)

    prefix = "FAC" if doc_type == "facture" else "BL"
    count = await db.documents.count_documents({"doc_type": doc_type})
    doc_number = f"{prefix}-{str(count + 1).zfill(3)}"

    items = body.get("items", [])
    subtotal = sum(i.get("quantity", 0) * i.get("unit_price", 0) for i in items)
    discount = body.get("discount", 0)
    total = subtotal - discount

    doc = {
        "doc_type": doc_type,
        "doc_number": doc_number,
        "client_name": body.get("client_name", ""),
        "client_phone": body.get("client_phone", ""),
        "client_email": body.get("client_email", ""),
        "client_address": body.get("client_address", ""),
        "items": items,
        "subtotal": subtotal,
        "discount": discount,
        "total": total,
        "notes": body.get("notes", ""),
        "created_by": user["_id"],
        "created_by_name": user.get("name", ""),
        "created_at": now.isoformat(),
    }
    result = await db.documents.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    await log_activity(user["_id"], "document_create", f"{prefix}-{doc_number}: {body.get('client_name','')}")
    return doc

@api_router.get("/documents/{doc_id}")
async def get_document(doc_id: str, request: Request):
    await get_current_user(request)
    doc = await db.documents.find_one({"_id": ObjectId(doc_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Document non trouvé")
    doc["_id"] = str(doc["_id"])
    return doc

@api_router.get("/documents/{doc_id}/pdf")
async def download_document_pdf(doc_id: str, request: Request, token: Optional[str] = None):
    # Accept token from query param for PDF download links
    if token:
        try:
            payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
            if payload.get("type") != "access":
                raise HTTPException(status_code=401, detail="Token invalide")
        except Exception:
            raise HTTPException(status_code=401, detail="Token invalide")
    else:
        await get_current_user(request)
    doc = await db.documents.find_one({"_id": ObjectId(doc_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Document non trouvé")

    company = await db.settings.find_one({"type": "company"}, {"_id": 0}) or DEFAULT_COMPANY
    buf = generate_pdf(doc, company)
    filename = f"{doc['doc_number']}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})

def generate_pdf(doc, company):
    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    is_facture = doc.get("doc_type") == "facture"
    title = "FACTURE" if is_facture else "BON DE LIVRAISON"

    # Header background
    c.setFillColor(colors.HexColor("#0A3D73"))
    c.rect(0, h - 90, w, 90, fill=True, stroke=False)

    # Company name
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 20)
    c.drawString(30, h - 40, company.get("name", "R2A Industrie"))
    c.setFont("Helvetica", 9)
    y_info = h - 55
    if company.get("address"):
        c.drawString(30, y_info, company["address"])
        y_info -= 12
    line2 = []
    if company.get("phone"):
        line2.append(f"Tél: {company['phone']}")
    if company.get("email"):
        line2.append(f"Email: {company['email']}")
    if line2:
        c.drawString(30, y_info, "  |  ".join(line2))
        y_info -= 12
    line3 = []
    if company.get("rc"):
        line3.append(f"RC: {company['rc']}")
    if company.get("nif"):
        line3.append(f"NIF: {company['nif']}")
    if company.get("nis"):
        line3.append(f"NIS: {company['nis']}")
    if company.get("ai"):
        line3.append(f"AI: {company['ai']}")
    if line3:
        c.drawString(30, y_info, "  |  ".join(line3))

    # Document title + number
    c.setFont("Helvetica-Bold", 16)
    c.drawRightString(w - 30, h - 40, title)
    c.setFont("Helvetica-Bold", 13)
    c.drawRightString(w - 30, h - 58, f"N° {doc.get('doc_number', '')}")
    c.setFont("Helvetica", 9)
    date_str = ""
    if doc.get("created_at"):
        try:
            dt = datetime.fromisoformat(doc["created_at"])
            date_str = dt.strftime("%d/%m/%Y")
        except Exception:
            date_str = doc["created_at"][:10]
    c.drawRightString(w - 30, h - 73, f"Date: {date_str}")

    # Client box
    y = h - 120
    c.setStrokeColor(colors.HexColor("#CBD5E1"))
    c.setFillColor(colors.HexColor("#F8FAFC"))
    c.roundRect(w - 260, y - 65, 230, 65, 4, fill=True, stroke=True)
    c.setFillColor(colors.HexColor("#0F172A"))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(w - 250, y - 15, "CLIENT")
    c.setFont("Helvetica", 10)
    c.drawString(w - 250, y - 30, doc.get("client_name", ""))
    c.setFont("Helvetica", 8)
    cl_y = y - 43
    if doc.get("client_phone"):
        c.drawString(w - 250, cl_y, f"Tél: {doc['client_phone']}")
        cl_y -= 11
    if doc.get("client_email"):
        c.drawString(w - 250, cl_y, doc["client_email"])
        cl_y -= 11
    if doc.get("client_address"):
        c.drawString(w - 250, cl_y, doc["client_address"])

    # Table header
    y = h - 210
    cols = [30, 60, 280, 340, 410, 480]
    headers = ["#", "Référence", "Désignation", "Qté", "Prix Unit.", "Total"]
    c.setFillColor(colors.HexColor("#0A3D73"))
    c.rect(25, y - 5, w - 50, 22, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    for i, header in enumerate(headers):
        if i >= 3:
            c.drawRightString(cols[i] + 70, y + 3, header)
        else:
            c.drawString(cols[i], y + 3, header)

    # Table rows
    y -= 20
    c.setFillColor(colors.HexColor("#0F172A"))
    items = doc.get("items", [])
    for idx, item in enumerate(items):
        if y < 100:
            c.showPage()
            y = h - 60
        c.setFont("Helvetica", 9)
        if idx % 2 == 0:
            c.setFillColor(colors.HexColor("#F8FAFC"))
            c.rect(25, y - 5, w - 50, 18, fill=True, stroke=False)
        c.setFillColor(colors.HexColor("#0F172A"))
        c.drawString(cols[0], y + 1, str(idx + 1))
        c.drawString(cols[1], y + 1, str(item.get("reference", ""))[:30])
        c.drawString(cols[2], y + 1, str(item.get("name", item.get("description", "")))[:35])
        c.drawRightString(cols[3] + 70, y + 1, str(item.get("quantity", 0)))
        price = item.get("unit_price", 0)
        line_total = item.get("quantity", 0) * price
        c.drawRightString(cols[4] + 70, y + 1, f"{price:,.0f} DZD")
        c.drawRightString(cols[5] + 70, y + 1, f"{line_total:,.0f} DZD")
        y -= 18

    # Totals
    y -= 15
    c.setStrokeColor(colors.HexColor("#CBD5E1"))
    c.line(w - 220, y + 5, w - 30, y + 5)
    c.setFont("Helvetica", 10)
    c.drawString(w - 210, y - 12, "Sous-total:")
    c.drawRightString(w - 35, y - 12, f"{doc.get('subtotal', 0):,.0f} DZD")
    if doc.get("discount", 0) > 0:
        c.drawString(w - 210, y - 28, "Remise:")
        c.drawRightString(w - 35, y - 28, f"-{doc['discount']:,.0f} DZD")
        y -= 16
    c.setFillColor(colors.HexColor("#0A3D73"))
    c.rect(w - 220, y - 40, 190, 22, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(w - 210, y - 34, "TOTAL:")
    c.drawRightString(w - 35, y - 34, f"{doc.get('total', 0):,.0f} DZD")

    # Notes
    if doc.get("notes"):
        c.setFillColor(colors.HexColor("#64748B"))
        c.setFont("Helvetica", 8)
        c.drawString(30, y - 60, f"Notes: {doc['notes']}")

    # Footer
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.setFont("Helvetica", 7)
    c.drawCentredString(w / 2, 25, f"{company.get('name', 'R2A Industrie')} - Document généré le {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}")

    c.save()
    buf.seek(0)
    return buf

# ============ MONTHLY REPORT & EMAIL ============

import resend
import asyncio

async def generate_report_data():
    """Generate full monthly report data."""
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()

    total_products = await db.products.count_documents({"archived": {"$ne": True}})
    out_of_stock = await db.products.count_documents({"archived": {"$ne": True}, "quantity": 0})
    low_stock_items = await db.products.find(
        {"archived": {"$ne": True}, "quantity": {"$gt": 0}, "$expr": {"$lte": ["$quantity", "$stock_minimum"]}},
        {"_id": 0, "reference": 1, "name": 1, "quantity": 1, "category": 1}
    ).limit(30).to_list(30)

    # Sales this month
    sales_agg = await db.sales.aggregate([
        {"$match": {"created_at": {"$gte": month_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    sales_total = sales_agg[0]["total"] if sales_agg else 0
    sales_count = sales_agg[0]["count"] if sales_agg else 0

    # Top sold products
    top_products = await db.sales.aggregate([
        {"$match": {"created_at": {"$gte": month_start}}},
        {"$unwind": "$items"},
        {"$group": {"_id": "$items.reference", "name": {"$first": "$items.name"}, "total_qty": {"$sum": "$items.quantity"}}},
        {"$sort": {"total_qty": -1}}, {"$limit": 10}
    ]).to_list(10)

    # New clients this month
    new_clients = await db.clients.count_documents({"created_at": {"$gte": month_start}, "archived": {"$ne": True}})
    total_clients = await db.clients.count_documents({"archived": {"$ne": True}})

    # User activity this month
    user_activity = await db.activity_logs.aggregate([
        {"$match": {"timestamp": {"$gte": month_start}}},
        {"$group": {"_id": "$user_name", "count": {"$sum": 1}, "actions": {"$push": "$action"}}},
        {"$sort": {"count": -1}}
    ]).to_list(20)

    # Recent important changes
    recent_logs = await db.activity_logs.find(
        {"timestamp": {"$gte": month_start}}, {"_id": 0}
    ).sort("timestamp", -1).limit(20).to_list(20)

    return {
        "month": now.strftime("%B %Y"),
        "generated_at": now.isoformat(),
        "total_products": total_products,
        "out_of_stock": out_of_stock,
        "low_stock_count": len(low_stock_items),
        "low_stock_items": low_stock_items,
        "sales_total": sales_total,
        "sales_count": sales_count,
        "top_products": top_products,
        "new_clients": new_clients,
        "total_clients": total_clients,
        "user_activity": user_activity,
        "recent_logs": recent_logs,
    }

def build_report_html(data):
    """Build HTML email from report data."""
    low_stock_rows = ""
    for p in data.get("low_stock_items", [])[:15]:
        low_stock_rows += f"<tr><td style='padding:6px 12px;border-bottom:1px solid #E2E8F0;font-family:monospace;color:#0A3D73'>{p['reference']}</td><td style='padding:6px 12px;border-bottom:1px solid #E2E8F0'>{p.get('name','')}</td><td style='padding:6px 12px;border-bottom:1px solid #E2E8F0;text-align:center;color:#EA580C;font-weight:bold'>{p['quantity']}</td></tr>"

    top_rows = ""
    for p in data.get("top_products", [])[:10]:
        top_rows += f"<tr><td style='padding:6px 12px;border-bottom:1px solid #E2E8F0;font-family:monospace'>{p['_id']}</td><td style='padding:6px 12px;border-bottom:1px solid #E2E8F0;text-align:center;font-weight:bold'>{p['total_qty']}</td></tr>"

    activity_rows = ""
    for u in data.get("user_activity", []):
        activity_rows += f"<tr><td style='padding:6px 12px;border-bottom:1px solid #E2E8F0;font-weight:500'>{u['_id'] or 'Système'}</td><td style='padding:6px 12px;border-bottom:1px solid #E2E8F0;text-align:center'>{u['count']}</td></tr>"

    top_section = ""
    if top_rows:
        top_section = '<h3 style="font-size:15px;margin:20px 0 8px;color:#0A3D73">Top articles vendus</h3><table style="width:100%;border-collapse:collapse"><tr style="background:#F1F5F9"><th style="padding:8px 12px;text-align:left;font-size:12px;text-transform:uppercase">Référence</th><th style="padding:8px 12px;text-align:center;font-size:12px;text-transform:uppercase">Qté vendue</th></tr>' + top_rows + '</table>'

    low_section = ""
    if low_stock_rows:
        low_section = '<h3 style="font-size:15px;margin:20px 0 8px;color:#EA580C">Articles en stock faible</h3><table style="width:100%;border-collapse:collapse"><tr style="background:#FFF7ED"><th style="padding:8px 12px;text-align:left;font-size:12px">Réf</th><th style="padding:8px 12px;text-align:left;font-size:12px">Nom</th><th style="padding:8px 12px;text-align:center;font-size:12px">Qté</th></tr>' + low_stock_rows + '</table>'

    activity_section = ""
    if activity_rows:
        activity_section = '<h3 style="font-size:15px;margin:20px 0 8px;color:#0A3D73">Activité utilisateurs</h3><table style="width:100%;border-collapse:collapse"><tr style="background:#F1F5F9"><th style="padding:8px 12px;text-align:left;font-size:12px">Utilisateur</th><th style="padding:8px 12px;text-align:center;font-size:12px">Actions</th></tr>' + activity_rows + '</table>'

    return f"""
    <div style="max-width:650px;margin:0 auto;font-family:Arial,sans-serif;color:#0F172A">
      <div style="background:#0A3D73;padding:24px 30px;border-radius:8px 8px 0 0">
        <h1 style="color:white;margin:0;font-size:22px">R2A Industrie - Rapport Mensuel</h1>
        <p style="color:rgba(255,255,255,0.7);margin:4px 0 0;font-size:14px">{data['month']}</p>
      </div>
      <div style="background:white;padding:24px 30px;border:1px solid #E2E8F0;border-top:none">
        <table style="width:100%;border-collapse:collapse;margin-bottom:24px">
          <tr>
            <td style="padding:16px;background:#EFF6FF;border-radius:6px;text-align:center;width:33%">
              <div style="font-size:28px;font-weight:bold;color:#0A3D73">{data['sales_count']}</div>
              <div style="font-size:12px;color:#64748B;text-transform:uppercase;margin-top:4px">Ventes</div>
            </td>
            <td style="width:8px"></td>
            <td style="padding:16px;background:#F0FDF4;border-radius:6px;text-align:center;width:33%">
              <div style="font-size:28px;font-weight:bold;color:#16A34A">{data['sales_total']:,.0f}</div>
              <div style="font-size:12px;color:#64748B;text-transform:uppercase;margin-top:4px">CA (DZD)</div>
            </td>
            <td style="width:8px"></td>
            <td style="padding:16px;background:#FEF2F2;border-radius:6px;text-align:center;width:33%">
              <div style="font-size:28px;font-weight:bold;color:#DC2626">{data['out_of_stock']}</div>
              <div style="font-size:12px;color:#64748B;text-transform:uppercase;margin-top:4px">Ruptures</div>
            </td>
          </tr>
        </table>
        <table style="width:100%;border-collapse:collapse;margin-bottom:20px">
          <tr>
            <td style="padding:10px 16px;background:#F8FAFC;border-radius:4px"><strong>Total articles :</strong> {data['total_products']}</td>
            <td style="padding:10px 16px;background:#F8FAFC;border-radius:4px"><strong>Stock faible :</strong> <span style="color:#EA580C">{data['low_stock_count']}</span></td>
          </tr>
          <tr><td style="height:8px" colspan="2"></td></tr>
          <tr>
            <td style="padding:10px 16px;background:#F8FAFC;border-radius:4px"><strong>Total clients :</strong> {data['total_clients']}</td>
            <td style="padding:10px 16px;background:#F8FAFC;border-radius:4px"><strong>Nouveaux clients :</strong> {data['new_clients']}</td>
          </tr>
        </table>
        {top_section}
        {low_section}
        {activity_section}
      </div>
      <div style="text-align:center;padding:16px;color:#94A3B8;font-size:11px">
        R2A Industrie - Rapport généré automatiquement
      </div>
    </div>"""

@api_router.get("/reports/monthly")
async def get_monthly_report(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    data = await generate_report_data()
    data["html"] = build_report_html(data)
    return data

@api_router.post("/reports/send-email")
async def send_monthly_report(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")

    api_key = os.environ.get("RESEND_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=400, detail="Clé API Resend non configurée. Ajoutez RESEND_API_KEY dans les paramètres.")

    resend.api_key = api_key
    recipient = os.environ.get("REPORT_RECIPIENT", "boukhalfarabah23@gmail.com")
    sender = os.environ.get("SENDER_EMAIL", "onboarding@resend.dev")

    data = await generate_report_data()
    html = build_report_html(data)

    try:
        result = await asyncio.to_thread(resend.Emails.send, {
            "from": sender,
            "to": [recipient],
            "subject": f"R2A Industrie - Rapport Mensuel {data['month']}",
            "html": html,
        })
        await log_activity(user["_id"], "report_sent", f"Rapport mensuel envoyé à {recipient}")
        return {"message": f"Rapport envoyé à {recipient}", "email_id": result.get("id", "")}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur d'envoi: {str(e)}")

# ============ ACTIVITY LOG ROUTES ============

@api_router.get("/activity")
async def list_activity(request: Request, page: int = 1, limit: int = 50, user_filter: Optional[str] = None, action_filter: Optional[str] = None):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    query = {}
    if user_filter:
        query["user_id"] = user_filter
    if action_filter:
        query["action"] = {"$regex": action_filter}
    total = await db.activity_logs.count_documents(query)
    cursor = db.activity_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip((page - 1) * limit).limit(limit)
    items = await cursor.to_list(limit)
    return {"items": items, "total": total, "page": page, "pages": math.ceil(total / limit) if limit > 0 else 0}

# ============ EXPORT ROUTES ============

@api_router.get("/export/products")
async def export_products_csv(request: Request, category: Optional[str] = None, token: Optional[str] = None):
    if token:
        try:
            jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        except Exception:
            raise HTTPException(status_code=401, detail="Token invalide")
    else:
        await get_current_user(request)
    query = {"archived": {"$ne": True}}
    if category:
        query["category"] = category
    cursor = db.products.find(query, {"_id": 0}).sort("reference", 1)
    products = await cursor.to_list(10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"
    headers = ["Référence", "Nom", "Catégorie", "Marque", "Quantité", "Stock Min", "Prix Achat", "Prix Vente", "Fournisseur", "Emplacement", "État", "Statut"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for p in products:
        ws.append([p.get("reference",""), p.get("name",""), p.get("category",""), p.get("brand",""), p.get("quantity",0), p.get("stock_minimum",0), p.get("purchase_price",0), p.get("sale_price",0), p.get("supplier",""), p.get("location",""), p.get("state",""), p.get("status","")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    cat_name = category or "tous"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=produits_{cat_name}.xlsx"})

@api_router.get("/export/sales")
async def export_sales_csv(request: Request, token: Optional[str] = None):
    if token:
        try:
            jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        except Exception:
            raise HTTPException(status_code=401, detail="Token invalide")
    else:
        await get_current_user(request)
    cursor = db.sales.find({}, {"_id": 0}).sort("created_at", -1)
    sales = await cursor.to_list(10000)
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["N° Vente", "Date", "Client", "Total", "Vendeur", "Articles"])
    for s in sales:
        items_str = " | ".join([f"{i.get('reference','')} x{i.get('quantity',0)}" for i in s.get("items", [])])
        writer.writerow([s.get("sale_number",""), s.get("created_at","")[:10], s.get("client_name",""), s.get("total_amount",0), s.get("sold_by_name",""), items_str])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=ventes_r2a.csv"})

# Bulk delete products
@api_router.post("/products/bulk-delete")
async def bulk_delete_products(request: Request):
    user = await get_current_user(request)
    body = await request.json()
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="Aucun produit sélectionné")
    now = datetime.now(timezone.utc).isoformat()
    result = await db.products.update_many(
        {"_id": {"$in": [ObjectId(i) for i in ids]}},
        {"$set": {"archived": True, "archived_at": now}}
    )
    await log_activity(user["_id"], "product_bulk_archive", f"{result.modified_count} produits archivés")
    return {"message": f"{result.modified_count} produits déplacés dans la corbeille"}

# ============ ALERTS ============

@api_router.get("/alerts")
async def get_alerts(request: Request):
    await get_current_user(request)
    low_stock = await db.products.find(
        {"archived": {"$ne": True}, "quantity": {"$gt": 0}, "$expr": {"$lte": ["$quantity", "$stock_minimum"]}},
        {"_id": 0, "reference": 1, "name": 1, "quantity": 1, "stock_minimum": 1, "category": 1}
    ).limit(20).to_list(20)
    out_of_stock = await db.products.find(
        {"archived": {"$ne": True}, "quantity": 0},
        {"_id": 0, "reference": 1, "name": 1, "category": 1}
    ).limit(20).to_list(20)
    return {"low_stock": low_stock, "out_of_stock": out_of_stock}

# ============ IMPORT EXCEL/CSV ============

import openpyxl
import csv
import io

CATEGORY_KEYWORDS = {
    "automate": ["automate"],
    "variateur": ["variateur"],
    "verin_pneumatique": ["verin_pneumatique", "verin pneumatique", "verin pneum"],
    "vapeur": ["vapeur"],
    "relais_securite": ["relais", "relais de sec", "relais_de_s"],
    "ecran_siemens": ["ecran", "ecran_siemens"],
    "hydrolique": ["hydrolique", "hydraulique", "hydraulic"],
    "pneumatique": ["pneumatique"],
    "encodeur_occasion": ["encodeur"],
    "instrument": ["instrument", "instrumentation"],
    "compteur": ["compteur"],
    "capteur": ["capteur"],
}

def detect_category_from_text(text: str) -> str:
    """Detect product category from filename or title."""
    text_lower = text.lower().replace("é", "e").replace("è", "e").replace("ê", "e")
    for cat_id, keywords in CATEGORY_KEYWORDS.items():
        for kw in keywords:
            if kw in text_lower:
                return cat_id
    return "automatisme"

def parse_excel_data(file_bytes: bytes, filename: str):
    """Parse Excel file and extract reference/brand/quantity rows."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []
    detected_category = detect_category_from_text(filename)
    title_text = ""

    for row in ws.iter_rows(values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        # Single value = title or sub-header
        if len(vals) == 1:
            val = str(vals[0]).strip()
            if not title_text and len(val) > 2:
                title_text = val
                cat = detect_category_from_text(val)
                if cat != "automatisme" or detected_category == "automatisme":
                    detected_category = cat
            continue

        # Data row: at least reference + brand
        ref = str(vals[0]).strip()
        brand = str(vals[1]).strip() if len(vals) > 1 else ""
        qty = 1

        # Try to parse quantity from 3rd column
        if len(vals) > 2:
            try:
                qty = int(float(vals[2]))
            except (ValueError, TypeError):
                qty = 1

        # Skip if reference looks like a brand name only (all uppercase, no numbers, short)
        if not ref or len(ref) < 2:
            continue

        # Skip rows that are just brand headers
        known_brands = ["OMRON", "SIEMENS", "SCHNEIDER", "FESTO", "SMC", "BOSCH", "REXROTH",
                        "ALLEN-BRADLEY", "TELEMECANIQUE", "DANFOSS", "LENZE", "SEW", "HITACHI",
                        "PARKER", "ABB", "MITSUBISHI", "FANUC", "YASKAWA", "BECKHOFF", "PILZ",
                        "WAGO", "IFM", "SICK", "JOUCOMATIC", "NUMATICS", "MARTONAIR"]
        if ref.upper() in known_brands and not brand:
            continue

        rows.append({
            "reference": ref,
            "brand": brand.upper() if brand else "",
            "quantity": max(qty, 0),
        })

    return rows, detected_category, title_text

def parse_csv_data(file_bytes: bytes, filename: str):
    """Parse CSV file."""
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=";")
    rows = []
    detected_category = detect_category_from_text(filename)

    for row_data in reader:
        vals = [v.strip() for v in row_data if v.strip()]
        if len(vals) < 2:
            continue
        ref = vals[0]
        brand = vals[1] if len(vals) > 1 else ""
        qty = 1
        if len(vals) > 2:
            try:
                qty = int(float(vals[2]))
            except (ValueError, TypeError):
                qty = 1
        if len(ref) >= 2:
            rows.append({"reference": ref, "brand": brand.upper(), "quantity": max(qty, 0)})

    return rows, detected_category, ""

@api_router.post("/import/preview")
async def import_preview(request: Request, file: UploadFile = File(...), category: str = Form("")):
    """Parse file and return preview data without importing."""
    await get_current_user(request)
    contents = await file.read()
    filename = file.filename or "unknown.xlsx"

    if filename.endswith(".csv"):
        rows, auto_category, title = parse_csv_data(contents, filename)
    else:
        rows, auto_category, title = parse_excel_data(contents, filename)

    chosen_category = category if category else auto_category

    # Check duplicates against existing DB
    existing_refs = set()
    cursor = db.products.find({"archived": {"$ne": True}}, {"reference": 1, "_id": 0})
    async for doc in cursor:
        existing_refs.add(doc.get("reference", ""))

    preview = []
    new_count = 0
    update_count = 0
    for row in rows:
        is_duplicate = row["reference"] in existing_refs
        if is_duplicate:
            update_count += 1
        else:
            new_count += 1
        preview.append({
            **row,
            "is_duplicate": is_duplicate,
            "action": "update" if is_duplicate else "create",
        })

    return {
        "filename": filename,
        "detected_category": auto_category,
        "chosen_category": chosen_category,
        "title": title,
        "total_rows": len(preview),
        "new_count": new_count,
        "update_count": update_count,
        "preview": preview[:100],  # First 100 for preview
        "all_data": preview,
    }

@api_router.post("/import/execute")
async def import_execute(request: Request):
    """Execute the import with validated data."""
    user = await get_current_user(request)
    body = await request.json()
    items = body.get("items", [])
    category = body.get("category", "automatisme")

    now = datetime.now(timezone.utc).isoformat()
    created = 0
    updated = 0
    errors = []

    for item in items:
        ref = str(item.get("reference", "")).strip()
        if not ref:
            errors.append({"reference": ref, "error": "Référence vide"})
            continue

        brand = item.get("brand", "")
        qty = item.get("quantity", 1)

        try:
            existing = await db.products.find_one({"reference": ref})
            if existing:
                # Update quantity and brand
                update_fields = {"updated_at": now}
                if brand:
                    update_fields["brand"] = brand
                if qty > 0:
                    update_fields["quantity"] = qty
                await db.products.update_one({"reference": ref}, {"$set": update_fields})
                updated += 1
            else:
                doc = {
                    "reference": ref,
                    "name": ref,
                    "category": category,
                    "quantity": qty,
                    "stock_minimum": 1,
                    "purchase_price": 0,
                    "sale_price": 0,
                    "supplier": "",
                    "location": "",
                    "brand": brand,
                    "description": "",
                    "state": "occasion",
                    "status": "en_stock" if qty > 0 else "rupture",
                    "archived": False,
                    "created_at": now,
                    "updated_at": now,
                }
                await db.products.insert_one(doc)
                created += 1
        except Exception as e:
            errors.append({"reference": ref, "error": str(e)})

    # Log import
    await log_activity(user["_id"], "import", f"Import {category}: {created} créés, {updated} mis à jour, {len(errors)} erreurs")

    # Save import history
    await db.import_history.insert_one({
        "filename": body.get("filename", ""),
        "category": category,
        "created": created,
        "updated": updated,
        "errors": len(errors),
        "total": len(items),
        "user_id": user["_id"],
        "timestamp": now,
    })

    return {
        "created": created,
        "updated": updated,
        "errors": errors[:50],
        "error_count": len(errors),
        "total": len(items),
    }

@api_router.get("/import/history")
async def import_history(request: Request):
    await get_current_user(request)
    cursor = db.import_history.find({}).sort("timestamp", -1).limit(20)
    items = []
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        items.append(doc)
    return {"items": items}

# ============ ACCOUNTING ROUTES ============

REVENUE_CATEGORIES = [
    {"id": "ventes", "label": "Ventes de produits"},
    {"id": "services", "label": "Services"},
    {"id": "autres_revenus", "label": "Autres revenus"},
]
EXPENSE_CATEGORIES = [
    {"id": "achats_stock", "label": "Achats de stock"},
    {"id": "frais_generaux", "label": "Frais généraux"},
    {"id": "salaires", "label": "Salaires"},
    {"id": "loyer", "label": "Loyer et charges"},
    {"id": "transport", "label": "Transport et livraison"},
    {"id": "marketing", "label": "Marketing"},
    {"id": "maintenance", "label": "Maintenance"},
    {"id": "autres_depenses", "label": "Autres dépenses"},
]

def serialize_doc(doc, extra_fields=None):
    """Generic MongoDB doc serializer."""
    result = {"id": str(doc["_id"])}
    for k, v in doc.items():
        if k == "_id":
            continue
        result[k] = v
    return result

# --- Accounting Dashboard ---
@api_router.get("/accounting/dashboard")
async def accounting_dashboard(request: Request, year: Optional[int] = None, month: Optional[int] = None):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")

    now = datetime.now(timezone.utc)
    target_year = year or now.year
    target_month = month or now.month

    # Current month range
    month_start = datetime(target_year, target_month, 1, tzinfo=timezone.utc).isoformat()
    if target_month == 12:
        month_end = datetime(target_year + 1, 1, 1, tzinfo=timezone.utc).isoformat()
    else:
        month_end = datetime(target_year, target_month + 1, 1, tzinfo=timezone.utc).isoformat()

    year_start = datetime(target_year, 1, 1, tzinfo=timezone.utc).isoformat()
    year_end = datetime(target_year + 1, 1, 1, tzinfo=timezone.utc).isoformat()

    # Monthly totals
    rev_month = await db.revenues.aggregate([
        {"$match": {"date": {"$gte": month_start, "$lt": month_end}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    exp_month = await db.expenses.aggregate([
        {"$match": {"date": {"$gte": month_start, "$lt": month_end}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}, "count": {"$sum": 1}}}
    ]).to_list(1)

    # Yearly totals
    rev_year = await db.revenues.aggregate([
        {"$match": {"date": {"$gte": year_start, "$lt": year_end}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    exp_year = await db.expenses.aggregate([
        {"$match": {"date": {"$gte": year_start, "$lt": year_end}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)

    # Monthly chart data (12 months)
    monthly_data = []
    for m in range(1, 13):
        ms = datetime(target_year, m, 1, tzinfo=timezone.utc).isoformat()
        me = datetime(target_year + 1, 1, 1, tzinfo=timezone.utc).isoformat() if m == 12 else datetime(target_year, m + 1, 1, tzinfo=timezone.utc).isoformat()
        r = await db.revenues.aggregate([{"$match": {"date": {"$gte": ms, "$lt": me}}}, {"$group": {"_id": None, "total": {"$sum": "$amount"}}}]).to_list(1)
        e = await db.expenses.aggregate([{"$match": {"date": {"$gte": ms, "$lt": me}}}, {"$group": {"_id": None, "total": {"$sum": "$amount"}}}]).to_list(1)
        monthly_data.append({
            "month": m,
            "revenue": r[0]["total"] if r else 0,
            "expense": e[0]["total"] if e else 0,
        })

    # Invoice stats
    invoices_pending = await db.invoices.count_documents({"status": "en_attente"})
    invoices_overdue = await db.invoices.count_documents({"status": "en_retard"})
    invoices_paid = await db.invoices.count_documents({"status": "payee"})

    # Revenue by category
    rev_by_cat = await db.revenues.aggregate([
        {"$match": {"date": {"$gte": month_start, "$lt": month_end}}},
        {"$group": {"_id": "$category", "total": {"$sum": "$amount"}}}
    ]).to_list(20)
    exp_by_cat = await db.expenses.aggregate([
        {"$match": {"date": {"$gte": month_start, "$lt": month_end}}},
        {"$group": {"_id": "$category", "total": {"$sum": "$amount"}}}
    ]).to_list(20)

    rev_m = rev_month[0]["total"] if rev_month else 0
    exp_m = exp_month[0]["total"] if exp_month else 0
    rev_y = rev_year[0]["total"] if rev_year else 0
    exp_y = exp_year[0]["total"] if exp_year else 0

    return {
        "month_revenue": rev_m,
        "month_expense": exp_m,
        "month_profit": rev_m - exp_m,
        "month_revenue_count": rev_month[0]["count"] if rev_month else 0,
        "month_expense_count": exp_month[0]["count"] if exp_month else 0,
        "year_revenue": rev_y,
        "year_expense": exp_y,
        "year_profit": rev_y - exp_y,
        "monthly_chart": monthly_data,
        "invoices_pending": invoices_pending,
        "invoices_overdue": invoices_overdue,
        "invoices_paid": invoices_paid,
        "revenue_by_category": {r["_id"]: r["total"] for r in rev_by_cat},
        "expense_by_category": {e["_id"]: e["total"] for e in exp_by_cat},
        "revenue_categories": REVENUE_CATEGORIES,
        "expense_categories": EXPENSE_CATEGORIES,
        "year": target_year,
        "month": target_month,
    }

# --- Revenues CRUD ---
@api_router.get("/accounting/revenues")
async def list_revenues(request: Request, page: int = 1, limit: int = 50, search: Optional[str] = None, category: Optional[str] = None):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    query = {}
    if category:
        query["category"] = category
    if search:
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query["$or"] = [{"description": pattern}, {"client_name": pattern}, {"invoice_ref": pattern}]
    total = await db.revenues.count_documents(query)
    cursor = db.revenues.find(query).sort("date", -1).skip((page - 1) * limit).limit(limit)
    items = [serialize_doc(doc) async for doc in cursor]
    return {"items": items, "total": total, "page": page, "pages": math.ceil(total / limit) if limit > 0 else 0}

@api_router.post("/accounting/revenues")
async def create_revenue(data: RevenueCreate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    now = datetime.now(timezone.utc).isoformat()
    doc = {**data.model_dump(), "date": data.date or now, "created_at": now, "created_by": user["_id"]}
    result = await db.revenues.insert_one(doc)
    await log_activity(user["_id"], "revenue_create", f"Revenu: {data.description} ({data.amount})")
    created = await db.revenues.find_one({"_id": result.inserted_id})
    return serialize_doc(created)

@api_router.put("/accounting/revenues/{rev_id}")
async def update_revenue(rev_id: str, data: RevenueUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.revenues.update_one({"_id": ObjectId(rev_id)}, {"$set": updates})
    await log_activity(user["_id"], "revenue_update", f"Revenu modifié: {rev_id}")
    updated = await db.revenues.find_one({"_id": ObjectId(rev_id)})
    if not updated:
        raise HTTPException(status_code=404, detail="Revenu non trouvé")
    return serialize_doc(updated)

@api_router.delete("/accounting/revenues/{rev_id}")
async def delete_revenue(rev_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    result = await db.revenues.delete_one({"_id": ObjectId(rev_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Revenu non trouvé")
    await log_activity(user["_id"], "revenue_delete", f"Revenu supprimé: {rev_id}")
    return {"message": "Revenu supprimé"}

# --- Expenses CRUD ---
@api_router.get("/accounting/expenses")
async def list_expenses(request: Request, page: int = 1, limit: int = 50, search: Optional[str] = None, category: Optional[str] = None):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    query = {}
    if category:
        query["category"] = category
    if search:
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query["$or"] = [{"description": pattern}, {"supplier_name": pattern}]
    total = await db.expenses.count_documents(query)
    cursor = db.expenses.find(query).sort("date", -1).skip((page - 1) * limit).limit(limit)
    items = [serialize_doc(doc) async for doc in cursor]
    return {"items": items, "total": total, "page": page, "pages": math.ceil(total / limit) if limit > 0 else 0}

@api_router.post("/accounting/expenses")
async def create_expense(data: ExpenseCreate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    now = datetime.now(timezone.utc).isoformat()
    doc = {**data.model_dump(), "date": data.date or now, "created_at": now, "created_by": user["_id"]}
    result = await db.expenses.insert_one(doc)
    await log_activity(user["_id"], "expense_create", f"Dépense: {data.description} ({data.amount})")
    created = await db.expenses.find_one({"_id": result.inserted_id})
    return serialize_doc(created)

@api_router.put("/accounting/expenses/{exp_id}")
async def update_expense(exp_id: str, data: ExpenseUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.expenses.update_one({"_id": ObjectId(exp_id)}, {"$set": updates})
    await log_activity(user["_id"], "expense_update", f"Dépense modifiée: {exp_id}")
    updated = await db.expenses.find_one({"_id": ObjectId(exp_id)})
    if not updated:
        raise HTTPException(status_code=404, detail="Dépense non trouvée")
    return serialize_doc(updated)

@api_router.delete("/accounting/expenses/{exp_id}")
async def delete_expense(exp_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    result = await db.expenses.delete_one({"_id": ObjectId(exp_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dépense non trouvée")
    await log_activity(user["_id"], "expense_delete", f"Dépense supprimée: {exp_id}")
    return {"message": "Dépense supprimée"}

# --- Invoices CRUD ---
@api_router.get("/accounting/invoices")
async def list_invoices(request: Request, page: int = 1, limit: int = 50, status: Optional[str] = None, search: Optional[str] = None):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    query = {}
    if status:
        query["status"] = status
    if search:
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query["$or"] = [{"client_name": pattern}, {"invoice_number": pattern}]
    total = await db.invoices.count_documents(query)
    cursor = db.invoices.find(query).sort("created_at", -1).skip((page - 1) * limit).limit(limit)
    items = [serialize_doc(doc) async for doc in cursor]
    return {"items": items, "total": total, "page": page, "pages": math.ceil(total / limit) if limit > 0 else 0}

@api_router.post("/accounting/invoices")
async def create_invoice(data: InvoiceCreate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    now = datetime.now(timezone.utc)
    # Auto-generate invoice number
    count = await db.invoices.count_documents({})
    inv_number = f"FAC-{now.strftime('%Y%m')}-{str(count + 1).zfill(4)}"

    subtotal = sum(item.get("quantity", 1) * item.get("unit_price", 0) for item in data.items)
    total = subtotal - data.discount

    doc = {
        "invoice_number": inv_number,
        "client_name": data.client_name,
        "client_email": data.client_email,
        "client_address": data.client_address,
        "items": data.items,
        "subtotal": subtotal,
        "discount": data.discount,
        "total": total,
        "notes": data.notes,
        "status": "en_attente",
        "due_date": data.due_date or (now + timedelta(days=30)).isoformat(),
        "created_at": now.isoformat(),
        "created_by": user["_id"],
    }
    result = await db.invoices.insert_one(doc)
    await log_activity(user["_id"], "invoice_create", f"Facture créée: {inv_number}")
    created = await db.invoices.find_one({"_id": result.inserted_id})
    return serialize_doc(created)

@api_router.put("/accounting/invoices/{inv_id}")
async def update_invoice(inv_id: str, data: InvoiceUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    existing = await db.invoices.find_one({"_id": ObjectId(inv_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if "items" in updates:
        subtotal = sum(i.get("quantity", 1) * i.get("unit_price", 0) for i in updates["items"])
        discount = updates.get("discount", existing.get("discount", 0))
        updates["subtotal"] = subtotal
        updates["total"] = subtotal - discount
    if "status" in updates and updates["status"] == "payee" and existing.get("status") != "payee":
        # Auto-create revenue when invoice is marked as paid
        rev_doc = {
            "description": f"Facture {existing.get('invoice_number', '')}",
            "amount": existing.get("total", 0),
            "category": "ventes",
            "client_name": existing.get("client_name", ""),
            "invoice_ref": existing.get("invoice_number", ""),
            "payment_method": "virement",
            "date": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user["_id"],
        }
        await db.revenues.insert_one(rev_doc)
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.invoices.update_one({"_id": ObjectId(inv_id)}, {"$set": updates})
    await log_activity(user["_id"], "invoice_update", f"Facture modifiée: {existing.get('invoice_number')}")
    updated = await db.invoices.find_one({"_id": ObjectId(inv_id)})
    return serialize_doc(updated)

@api_router.delete("/accounting/invoices/{inv_id}")
async def delete_invoice(inv_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    result = await db.invoices.delete_one({"_id": ObjectId(inv_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    await log_activity(user["_id"], "invoice_delete", f"Facture supprimée: {inv_id}")
    return {"message": "Facture supprimée"}

# ============ ACTIVITY LOG ============

async def log_activity(user_id: str, action: str, details: str, old_value: str = "", new_value: str = ""):
    # Get user name
    u = await db.users.find_one({"_id": ObjectId(user_id)}, {"name": 1}) if ObjectId.is_valid(user_id) else None
    await db.activity_logs.insert_one({
        "user_id": user_id,
        "user_name": u.get("name", "") if u else "",
        "action": action,
        "details": details,
        "old_value": old_value,
        "new_value": new_value,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })

# ============ STARTUP ============

@app.on_event("startup")
async def startup():
    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.products.create_index("reference", unique=True)
    await db.products.create_index("category")
    await db.products.create_index([("name", 1), ("reference", 1), ("brand", 1)])
    await db.products.create_index("archived")
    await db.clients.create_index("phone")
    await db.login_attempts.create_index("identifier")
    await db.activity_logs.create_index("timestamp")
    await db.revenues.create_index("date")
    await db.revenues.create_index("category")
    await db.expenses.create_index("date")
    await db.expenses.create_index("category")
    await db.invoices.create_index("invoice_number", unique=True)
    await db.invoices.create_index("status")
    await db.sales.create_index("created_at")
    await db.sales.create_index("client_id")
    await db.clients.create_index("name")
    await db.documents.create_index("doc_number", unique=True)
    await db.documents.create_index("doc_type")

    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@r2a-industrie.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin123!")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Administrateur",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin créé: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}}
        )
        logger.info(f"Admin mot de passe mis à jour: {admin_email}")

    # Write test credentials
    os.makedirs("/app/memory", exist_ok=True)
    with open("/app/memory/test_credentials.md", "w") as f:
        f.write("# Test Credentials\n\n")
        f.write(f"## Admin\n- Email: {admin_email}\n- Password: {admin_password}\n- Role: admin\n\n")
        f.write("## Auth Endpoints\n- POST /api/auth/login\n- POST /api/auth/register\n- GET /api/auth/me\n- POST /api/auth/logout\n- POST /api/auth/refresh\n\n")
        f.write("## Dashboard Endpoints\n- GET /api/dashboard/stats\n\n")
        f.write("## Products Endpoints\n- GET /api/products?category=X&search=X&page=1&limit=50&archived=false\n")
        f.write("- POST /api/products\n- GET /api/products/{id}\n- PUT /api/products/{id}\n- DELETE /api/products/{id} (soft delete)\n")
        f.write("- POST /api/products/{id}/restore\n- DELETE /api/products/{id}/permanent (admin only)\n")

    logger.info("Application R2A Industrie démarrée")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# Include router
app.include_router(api_router)

# CORS - must allow the frontend origin for cookies
frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
allowed_origins = [frontend_url]
# Also allow CORS_ORIGINS if set
cors_origins = os.environ.get("CORS_ORIGINS", "")
if cors_origins and cors_origins != "*":
    allowed_origins.extend([o.strip() for o in cors_origins.split(",") if o.strip()])

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
